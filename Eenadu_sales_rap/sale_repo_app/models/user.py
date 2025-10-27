from odoo import models, api, fields, _
from datetime import datetime, timedelta
import secrets
from odoo.exceptions import AccessDenied, ValidationError
from odoo.fields import Many2one
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
import base64
from odoo.exceptions import UserError
from io import BytesIO
from odoo.http import request
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from odoo.tools import round


class Users(models.Model):
    _inherit = 'res.users'

    role = fields.Selection(
        [('agent', 'Staff'),
         ('Office_staff', 'Office staff'),
         ('unit_manager', 'unit manager'),
         ('segment_incharge', 'segment incharge'),
         ('circulation_incharge', 'circulation incharge'),
         ('region_head', 'region head'),
         ('circulation_head', 'circulation head'),('admin','admin')],
        string="Role", required=True, default="Office_staff"
    )
    unit_name_ids = fields.One2many('unit.name','unit_name_id')
    status = fields.Selection(
        [('un_activ', 'Waiting For Approve'),
         ('active', 'Approved'),
      ],
        string="Role", required=True, default="un_activ"
    )

    def waiting_for_approve(self):
        self.ensure_one()
        self.status = 'un_activ'

    def approved_staff(self):
        self.ensure_one()
        self.status = 'active'

    pin_location_ids = fields.Many2many('pin.location','user_id')
    present_pin_id = fields.Many2one("pin.location")
    aadhar_number = fields.Char(string="Aadhar")
    root_name_id = fields.Many2one('root.map', string="Root Map")
    pan_number = fields.Char(string="PAN")
    state = fields.Char(string="state")
    phone = fields.Char(string="phone")
    user_id = fields.Integer(string="User ID")
    unit_name = fields.Char(string="Unit Name")
    create_uid = fields.Many2one(string="Created By",readonly=0)
    api_token = fields.Char(string="API Token", readonly=True)
    token_expiry = fields.Datetime(string="Token Expiry")
    aadhar_base64 = fields.Binary(string="Aadhar image")
    Pan_base64 = fields.Binary(string="Pan image")
    target = fields.Char(string="Target")
    edit_boll = fields.Boolean(string="Edit Allowed", compute="_compute_sale_user_readonly", store=False)
    # created_by = fields.Many2one('res.users', string="Created By", related="create_uid", store=True, readonly=False)

    created_by = fields.Many2one(
        'res.users',
        string="Created By",
        related="create_uid",
        store=True,
        readonly=False,  # make it editable
        inverse="_inverse_created_by"
    )

    @api.constrains('aadhar_number')
    def _check_aadhar_number(self):
        for user in self:
            if user.aadhar_number and not re.fullmatch(r'\d{12}', user.aadhar_number):
                raise ValidationError(_("Aadhar number must be exactly 12 digits and numeric only."))

    @api.constrains('phone')
    def _check_phone_number(self):
        for user in self:
            if user.phone and not re.fullmatch(r'\d{10}', user.phone):
                raise ValidationError(_("Phone number must be exactly 10 digits and numeric only."))

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            creating_user = self.env.user

            # 🛡️ Ensure password is either a string or not set
            password = vals.get('password')
            if isinstance(password, bool):  # e.g., True/False - INVALID
                raise ValidationError(_("Invalid password value. Must be a string."))

            # Auto-assign unit_name from creator
            if creating_user.unit_name and not vals.get('unit_name'):
                vals['unit_name'] = creating_user.unit_name

        users = super().create(vals_list)

        # Assign role-based groups
        for user in users:
            user._update_user_group_by_role()

        return users

    def _inverse_created_by(self):
        for rec in self:
            if rec.created_by:
                rec.write({'create_uid': rec.created_by.id})

    @api.depends_context('uid')
    def _compute_sale_user_readonly(self):
        current_user_in_group = self.env.user.has_group('sale_repo_app.circulation_incharge_group')
        for rec in self:
            rec.edit_boll = current_user_in_group

    def create_record(self):
        """
        Saves the record automatically and returns a success notification.
        """
        self.ensure_one()

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': "Success!",
                'message': "User created successfully.",
                'type': 'success',
                'sticky': False,
            }
        }

    def generate_token(self):
        """ Generate a unique API token and set an expiration time. """
        self.api_token = secrets.token_hex(32)  # Generates a unique 32-character token
        self.token_expiry = fields.Datetime.now() + timedelta(hours=10)  # Expires in 1 hour
        self.sudo().write({'api_token': self.api_token, 'token_expiry': self.token_expiry})  # Save token

    def authenticate_by_token(self, token):
        """ Validate the user using the API token. """
        user = self.sudo().search([('api_token', '=', token)], limit=1)
        if not user:
            raise AccessDenied(_("Invalid token!"))

        if user.token_expiry and user.token_expiry < fields.Datetime.now():
            raise AccessDenied(_("Token has expired! Please log in again."))

        return user

    def clear_token(self):
        """ Clear the token when the user logs out. """
        self.sudo().write({'api_token': False, 'token_expiry': False})


    ROLE_GROUP_MAPPING = {
        'admin': 'sale_repo_app.admin_group',
        'circulation_head': 'sale_repo_app.circulation_head_group',
        'region_head': 'sale_repo_app.region_head_group',
        'unit_manager': 'sale_repo_app.unit_manager_group',
        'segment_incharge': 'sale_repo_app.segment_incharge_group',
        'circulation_incharge': 'sale_repo_app.circulation_incharge_group',
        'Office_staff': 'sale_repo_app.office_staff_group',
        'agent': 'sale_repo_app.agent_group',
    }

    # --- Create override to assign group by role ---
    @api.model_create_multi
    def create(self, vals_list):
        users = super().create(vals_list)
        for user in users:
            user._update_user_group_by_role()
        return users

    # --- Write override to reassign group if role changes ---
    def write(self, vals):
        res = super().write(vals)
        if 'role' in vals:
            for user in self:
                user._update_user_group_by_role()
        return res

    # --- Core logic: remove all mapped groups and assign only current role group ---
    def _update_user_group_by_role(self):
        if not self.role:
            return

        group_ids = []
        for group_xml_id in self.ROLE_GROUP_MAPPING.values():
            try:
                group = self.env.ref(group_xml_id, raise_if_not_found=False)
                if group:
                    group_ids.append(group.id)
            except Exception:
                continue

        # Remove all role-based groups
        self.groups_id = [(3, gid) for gid in group_ids]

        # Assign new group based on selected role
        new_group_xml_id = self.ROLE_GROUP_MAPPING.get(self.role)
        if new_group_xml_id:
            new_group = self.env.ref(new_group_xml_id, raise_if_not_found=False)
            if new_group:
                self.groups_id = [(4, new_group.id)]

    @api.model
    def fields_get(self, allfields=None, attributes=None):
        fields_res = super().fields_get(allfields=allfields, attributes=attributes)
        if 'role' in fields_res:
            if not (self.env.user.has_group('sale_repo_app.region_head_group') or
                    self.env.user.has_group('sale_repo_app.circulation_head_group')):
                fields_res['role']['readonly'] = True
        return fields_res

    @api.model
    def fields_view_get(self, view_id=None, view_type='form', toolbar=False, submenu=False):
        res = super().fields_view_get(view_id=view_id, view_type=view_type,
                                      toolbar=toolbar, submenu=submenu)

        if view_type == 'form':
            doc = etree.XML(res['arch'])
            role_field = doc.xpath("//field[@name='role']")
            if role_field:
                # Example: only show 'role' if current user is region_head or circulation_head
                if not self.env.user.has_group('sale_repo_app.region_head_group') and \
                        not self.env.user.has_group('sale_repo_app.circulation_head_group'):
                    for node in role_field:
                        node.set('modifiers', '{"invisible": true}')
                        node.set('readonly', "1")

            res['arch'] = etree.tostring(doc, encoding='unicode')
        return res

class unit_names(models.Model):
    _name = 'unit.name'

    name = fields.Char(string="Name")
    unit_name_id = fields.Many2one('res.users')


from odoo import models, api, fields, _
from datetime import datetime, timedelta
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from io import BytesIO
import base64
from odoo.http import request
from odoo.exceptions import UserError, ValidationError
from openpyxl.utils import get_column_letter
import logging

_logger = logging.getLogger(__name__)


class UsersWizard(models.TransientModel):
    _name = "users.wizard"
    _description = "Users Wizard"

    # For creating a new res.users
    name = fields.Char(string="Name")
    login = fields.Char(string="Login")
    password = fields.Char(string="Password")

    # Custom fields
    role = fields.Selection(
        [
            ('agent', 'Staff'),
            ('Office_staff', 'Office staff'),
            ('unit_manager', 'Unit manager'),
            ('segment_incharge', 'Segment incharge'),
            ('circulation_incharge', 'Circulation incharge'),
            ('region_head', 'Region head'),
        ],
        string="Role",
    )
    status = fields.Selection(
        [('un_activ', 'Waiting For Approve'),
         ('active', 'Approved')],
        string="Status",
        default="un_activ",
        required=True,
    )
    unit_name_id = fields.Many2one('unit.name', string="Unit Name")
    aadhar_number = fields.Char(string="Aadhar")
    pan_number = fields.Char(string="PAN")
    phone = fields.Char(string="Phone")
    state = fields.Char(string="State")

    # for Customerforms xl report
    period_type = fields.Selection([
        ('day', 'Day Wise'),
        ('week', 'Week Wise'),
        ('month', 'Month Wise'),
        ('year', 'Yearly'),
        ('custom', 'Custom Range'),
        ('total', 'Total')
    ], string="Period Type", default='total', required=True)

    # Common field
    selected_year = fields.Integer(string="Year", default=lambda self: datetime.now().year)

    # Day wise
    selected_day = fields.Date(string="Select Day")

    # Week wise
    selected_week = fields.Integer(string="Week Number (1-53)", default=1)

    # Month wise
    month_selection = fields.Selection([
        ('1', 'January'),
        ('2', 'February'),
        ('3', 'March'),
        ('4', 'April'),
        ('5', 'May'),
        ('6', 'June'),
        ('7', 'July'),
        ('8', 'August'),
        ('9', 'September'),
        ('10', 'October'),
        ('11', 'November'),
        ('12', 'December')
    ], string="Month")

    # Manual dates for custom
    start_date = fields.Date(string="Start Date")
    end_date = fields.Date(string="End Date")

    dummy_file = fields.Binary("Excel File", readonly=True, attachment=True)
    dummy_file_name = fields.Char("File Name")
    unit_selection = fields.Selection([("HYD", "HYD"), ("warangal", "warangal"), ("All", "All")],
                                      default=lambda self: self._default_unit_selection())
    current_user_role = fields.Char(string="Current User Role", compute="_compute_current_user_role")

    # HTML Summary Field for Customer Forms Preview
    customer_html_summary = fields.Html(string="Customer Forms Summary", compute="_compute_customer_html_summary",
                                        sanitize=False)

    @api.depends('period_type')  # Dependency can be on any field; it's just to trigger recompute if needed
    def _compute_current_user_role(self):
        for rec in self:
            rec.current_user_role = self.env.user.role  # Fetches the logged-in user's role from res.users

    @api.model
    def _default_unit_selection(self):
        user = self.env.user
        if user.role == "circulation_incharge":
            return user.unit_name
        return "All"

    @api.depends('start_date', 'end_date', 'period_type')
    def _compute_customer_html_summary(self):
        for rec in self:
            if not rec.start_date or not rec.end_date:
                rec.customer_html_summary = '<p class="text-muted">No dates selected. Please select a valid period to view the summary.</p>'
                continue

            try:
                start_date, end_date = rec.start_date, rec.end_date
                user = self.env.user
                unit_names = []
                if user.role == "region_head":
                    unit_names = [u.name for u in user.unit_name_ids] if user.unit_name_ids else []
                else:
                    unit_name = getattr(user, 'unit_name', False)
                    if unit_name:
                        unit_names = [unit_name]

                if not unit_names:
                    rec.customer_html_summary = '<div class="alert alert-warning"><i class="fa fa-exclamation-triangle me-2"></i>No unit assigned to the user.</div>'
                    continue

                # Collect data for all units
                unit_data = []
                for unit in unit_names:
                    stats = self.env['customer.form'].get_customer_stats(
                        start_date=start_date,
                        end_date=end_date,
                        unit_name=unit
                    )
                    forms = stats["forms"]  # actual recordset

                    if not forms:
                        unit_data.append({
                            'unit': unit,
                            'man_days': 0,
                            'total_forms': 0,
                            'sport_count': 0,
                            'first_count': 0,
                            'aug': 0.0
                        })
                        continue

                    # Calculate man_days: number of unique (agent, date) pairs
                    unique_agent_days = set((f.agent_login, f.date) for f in forms if f.agent_login and f.date)
                    man_days = len(unique_agent_days)

                    # SPOT and 1st counts
                    first_count = sum(1 for f in forms if f.Start_Circulating and f.Start_Circulating[-2:] == "01")
                    sport_count = sum(1 for f in forms if f.Start_Circulating and f.Start_Circulating[-2:] != "01")
                    total_forms = stats["total_forms"]

                    # AVG = total_forms / man_days
                    aug = total_forms / man_days if man_days > 0 else 0.0

                    unit_data.append({
                        'unit': unit,
                        'man_days': man_days,
                        'total_forms': total_forms,
                        'sport_count': sport_count,
                        'first_count': first_count,
                        'aug': aug
                    })

                # Sort by total_forms descending
                unit_data.sort(key=lambda x: x['total_forms'], reverse=True)

                # Generate HTML table
                html = f"""
                <div class="card shadow-sm mt-3">
                    <div class="card-header bg-primary text-white">
                        <h6 class="mb-0">
                            <i class="fa fa-table me-2"></i>
                            Customer Forms Summary for {start_date.strftime('%d/%m/%Y')} to {end_date.strftime('%d/%m/%Y')}
                            <span class="badge bg-light text-dark ms-2">{', '.join(unit_names)}</span>
                        </h6>
                    </div>
                    <div class="card-body p-0">
                        <div class="table-responsive">
                            <table class="table table-hover mb-0">
                                <thead class="table-light">
                                    <tr>
                                        <th>SN</th>
                                        <th>Unit Name</th>
                                        <th>Man Days</th>
                                        <th>CPS (Total Forms)</th>
                                        <th>Spot</th>
                                        <th>1st</th>
                                        <th>AVG</th>
                                    </tr>
                                </thead>
                                <tbody>
                """

                for i, data in enumerate(unit_data, 1):
                    html += f"""
                                    <tr>
                                        <td>{i}</td>
                                        <td><strong>{data['unit']}</strong></td>
                                        <td><span class="badge bg-success">{data['man_days']}</span></td>
                                        <td><span class="badge bg-info">{data['total_forms']}</span></td>
                                        <td>{data['sport_count']}</td>
                                        <td>{data['first_count']}</td>
                                        <td>{data['aug']:.2f}</td>
                                    </tr>
                    """

                # Grand totals
                total_units = len([d for d in unit_data if d['total_forms'] > 0])  # Only units with data
                grand_total_forms = sum(d['total_forms'] for d in unit_data)
                grand_man_days = sum(d['man_days'] for d in unit_data)
                grand_avg = grand_total_forms / grand_man_days if grand_man_days > 0 else 0.0

                html += f"""
                                </tbody>
                            </table>
                        </div>
                        <div class="card-footer bg-light">
                            <div class="row">
                                <div class="col-md-3">
                                    <strong>Total Units: {total_units}</strong>
                                </div>
                                <div class="col-md-3">
                                    <strong>Total Man Days: {grand_man_days}</strong>
                                </div>
                                <div class="col-md-3">
                                    <strong>Total Forms: {grand_total_forms}</strong>
                                </div>
                                <div class="col-md-3">
                                    <strong>Overall AVG: {grand_avg:.2f}</strong>
                                </div>
                            </div>
                            {'<p class="text-muted small mb-0">Note: Values are based on available data in the selected period. If zeros appear, no forms were recorded.</p>' if grand_total_forms == 0 else ''}
                        </div>
                    </div>
                </div>
                """

                rec.customer_html_summary = html

            except Exception as e:
                _logger.error("Error in _compute_customer_html_summary: %s", e)
                rec.customer_html_summary = f'<div class="alert alert-danger"><i class="fa fa-exclamation-triangle me-2"></i>Error loading summary: {str(e)}</div>'

    @api.onchange('period_type')
    def _onchange_period_type(self):
        today = fields.Date.today()
        self.start_date = False
        self.end_date = False
        if self.period_type == 'total':
            self.selected_year = False
            self.selected_day = False
            self.selected_week = 1
            self.month_selection = False
            return
        elif self.period_type == 'day':
            self.selected_day = today
            self._onchange_selected_day()
            self.selected_year = False
            self.selected_week = 1
            self.month_selection = False
            return
        elif self.period_type == 'week':
            self.selected_week = ((today - datetime(today.year, 1, 1).date()).days // 7) + 1
            self._onchange_week()
            self.selected_year = today.year
            self.selected_day = False
            self.month_selection = False
            return
        elif self.period_type == 'month':
            self.month_selection = str(today.month)
            self._onchange_month()
            self.selected_year = today.year
            self.selected_day = False
            self.selected_week = 1
            return
        elif self.period_type == 'year':
            self.selected_year = today.year
            self._onchange_year()
            self.selected_day = False
            self.selected_week = 1
            self.month_selection = False
            return
        elif self.period_type == 'custom':
            # User will set manually
            pass

    @api.onchange('selected_day')
    def _onchange_selected_day(self):
        if self.period_type == 'day' and self.selected_day:
            self.start_date = self.selected_day
            self.end_date = self.selected_day

    @api.onchange('selected_week', 'selected_year')
    def _onchange_week(self):
        if self.period_type == 'week' and self.selected_week and self.selected_year:
            year = self.selected_year
            week = self.selected_week
            # First Monday of the year: Monday of the week containing Jan 4
            jan4 = datetime(year, 1, 4).date()
            first_monday = jan4 - timedelta(days=jan4.weekday())
            # Start of selected week
            start = first_monday + timedelta(weeks=week - 1)
            end = start + timedelta(days=6)
            self.start_date = start
            self.end_date = end

    @api.onchange('month_selection', 'selected_year')
    def _onchange_month(self):
        if self.period_type == 'month' and self.month_selection and self.selected_year:
            month = int(self.month_selection)
            year = self.selected_year
            start = datetime(year, month, 1).date()
            if month == 12:
                end = datetime(year + 1, 1, 1).date() - timedelta(days=1)
            else:
                end = datetime(year, month + 1, 1).date() - timedelta(days=1)
            self.start_date = start
            self.end_date = end

    @api.onchange('selected_year')
    def _onchange_year(self):
        if self.period_type == 'year' and self.selected_year:
            year = self.selected_year
            start = datetime(year, 1, 1).date()
            end = datetime(year, 12, 31).date()
            self.start_date = start
            self.end_date = end

    @api.constrains('selected_week')
    def _check_week(self):
        for rec in self:
            if rec.period_type == 'week' and (rec.selected_week < 1 or rec.selected_week > 53):
                raise ValidationError("Week number must be between 1 and 53.")

    @api.constrains('start_date', 'end_date')
    def _check_dates(self):
        for rec in self:
            if rec.period_type == 'custom' and rec.start_date and rec.end_date and rec.start_date > rec.end_date:
                raise ValidationError("Start date must be before or equal to end date.")

    @api.constrains('aadhar_number')
    def _check_aadhar_number(self):
        for user in self:
            if user.aadhar_number and not re.fullmatch(r'\d{12}', user.aadhar_number):
                raise ValidationError("Aadhar number must be exactly 12 digits and numeric only.")

    @api.constrains('phone')
    def _check_phone_number(self):
        for user in self:
            if user.phone and not re.fullmatch(r'\d{10}', user.phone):
                raise ValidationError("Phone number must be exactly 10 digits and numeric only.")

    def _get_report_dates(self):
        """Helper method to get start and end dates, calculating if necessary."""
        today = fields.Date.today()

        if self.period_type == 'total':
            # ✅ Set start date as September 1 of current year and end date as today
            start = datetime(today.year, 9, 1).date()
            end = today
            return start, end

        if self.start_date and self.end_date:
            return self.start_date, self.end_date

        if self.period_type == 'day':
            if self.selected_day:
                return self.selected_day, self.selected_day
            else:
                return today, today

        elif self.period_type == 'week':
            if self.selected_year and self.selected_week:
                year = self.selected_year
                week = self.selected_week
                jan4 = datetime(year, 1, 4).date()
                first_monday = jan4 - timedelta(days=jan4.weekday())
                start = first_monday + timedelta(weeks=week - 1)
                end = start + timedelta(days=6)
                return start, end
            else:
                days_to_monday = today.weekday()
                start = today - timedelta(days=days_to_monday)
                end = start + timedelta(days=6)
                return start, end

        elif self.period_type == 'month':
            if self.selected_year and self.month_selection:
                month = int(self.month_selection)
                year = self.selected_year
                start = datetime(year, month, 1).date()
                if month == 12:
                    end = datetime(year + 1, 1, 1).date() - timedelta(days=1)
                else:
                    end = datetime(year, month + 1, 1).date() - timedelta(days=1)
                return start, end
            else:
                start = today.replace(day=1)
                next_month = start.replace(month=start.month % 12 + 1)
                if next_month.month != 1:
                    end = next_month.replace(day=1) - timedelta(days=1)
                else:
                    end = next_month.replace(year=next_month.year - 1, month=12, day=1) - timedelta(days=1)
                return start, end

        elif self.period_type == 'year':
            if self.selected_year:
                year = self.selected_year
                start = datetime(year, 1, 1).date()
                end = datetime(year, 12, 31).date()
                return start, end
            else:
                year = today.year
                start = datetime(year, 1, 1).date()
                end = datetime(year, 12, 31).date()
                return start, end

        elif self.period_type == 'custom':
            return self.start_date, self.end_date

        return False, False

    def action_create_user(self):
        self.ensure_one()

        vals = {
            'name': self.name,
            'login': self.login,
            'password': self.password,
            'role': self.role,
            'status': self.status,
            'unit_name': self.unit_name_id.name if self.unit_name_id else False,
            'aadhar_number': self.aadhar_number,
            'pan_number': self.pan_number,
            'phone': self.phone,
            'state': self.state,
        }

        # Create the user with sudo
        user = self.env['res.users'].sudo().create(vals)

        # Clear the wizard record
        self.sudo().unlink()

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Success',
                'message': f'User {user.name} created successfully!',
                'type': 'success',
                'sticky': False,
                'close': {
                    'type': 'ir.actions.act_window',
                    'res_model': 'res.users',
                    'view_mode': 'tree,form',
                    'target': 'current',
                }
            }
        }

    def download_xl_report(self):
        start_date, end_date = self._get_report_dates()
        # if not start_date or not end_date:
        #     raise UserError("Please select valid period details.")

        for_Dates = f"{start_date} -- {end_date}"

        wb = Workbook()
        ws = wb.active
        ws.title = "Staff Analysis"

        # First row headings
        ws["A1"].value = "SN"
        ws["B1"].value = "UNIT NAME"

        # Merge for date range heading
        ws.merge_cells("C1:G1")
        main_heading_cell = ws["C1"]
        main_heading_cell.value = for_Dates
        main_heading_cell.alignment = Alignment(horizontal="center", vertical="center")
        main_heading_cell.font = Font(bold=True, size=12)

        # Merge for "PROMOTERES"
        ws.merge_cells("C2:G2")
        main_heading_cell = ws["C2"]
        main_heading_cell.value = "PROMOTERES"
        main_heading_cell.alignment = Alignment(horizontal="center", vertical="center")
        main_heading_cell.font = Font(bold=True, size=11)

        # Sub-headings in row 3
        ws["C3"].value = "MAN DAYS"
        ws["D3"].value = "CPS"
        ws["E3"].value = "SPOT"
        ws["F3"].value = "1st"
        ws["G3"].value = "AVG"

        # Make headers bold + bigger + centered
        header_cells = ["A1", "B1", "C3", "D3", "E3", "F3", "G3"]
        for cell_ref in header_cells:
            cell = ws[cell_ref]
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Set column widths
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 20
        for col in ["C", "D", "E", "F", "G"]:
            ws.column_dimensions[col].width = 12

        # Center align ALL cells (headers + data)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        user = request.env.user
        unit_names = []
        if user.role == "region_head":
            if not user.exists():
                return {"status": 404, "message": "User not found"}
            for i in user.unit_name_ids:
                unit_names.append(i.name)
        else:
            unit_name = user.unit_name
            if unit_name:
                unit_names = [unit_name]

        if not unit_names:
            raise UserError("No unit assigned to the user.")

        # Collect data for all units
        unit_data = []
        for unit in unit_names:
            stats = self.env['customer.form'].get_customer_stats(
                start_date=start_date,
                end_date=end_date,
                unit_name=unit
            )
            forms = stats["forms"]  # actual recordset

            if not forms:
                unit_data.append({
                    'unit': unit,
                    'man_days': 0,
                    'total_forms': 0,
                    'sport_count': 0,
                    'first_count': 0,
                    'aug': 0.0
                })
                continue

            # Calculate man_days: number of unique (agent, date) pairs
            unique_agent_days = set((f.agent_login, f.date) for f in forms if f.agent_login and f.date)
            man_days = len(unique_agent_days)

            # SPOT and 1st counts
            first_count = sum(1 for f in forms if f.Start_Circulating and f.Start_Circulating[-2:] == "01")
            sport_count = sum(1 for f in forms if f.Start_Circulating and f.Start_Circulating[-2:] != "01")
            total_forms = stats["total_forms"]

            # AVG = total_forms / man_days
            aug = total_forms / man_days if man_days > 0 else 0.0

            unit_data.append({
                'unit': unit,
                'man_days': man_days,
                'total_forms': total_forms,
                'sport_count': sport_count,
                'first_count': first_count,
                'aug': aug
            })

        # Sort by total_forms descending (more count at top, less at bottom)
        unit_data.sort(key=lambda x: x['total_forms'], reverse=True)

        # Fill data rows
        for i, data in enumerate(unit_data, start=1):
            num = 3 + i
            for j in range(1, 8):
                col_letter = chr(64 + j)
                cell = f"{col_letter}{num}"
                if j == 1:
                    ws[cell].value = str(i)  # SN
                elif j == 2:
                    ws[cell].value = data['unit']  # Unit Name
                elif j == 3:
                    ws[cell].value = data['man_days']  # MAN DAYS
                elif j == 4:
                    ws[cell].value = data['total_forms']  # CPS (Total Forms)
                elif j == 5:
                    ws[cell].value = data['sport_count']  # SPOT
                elif j == 6:
                    ws[cell].value = data['first_count']  # 1st
                elif j == 7:
                    ws[cell].value = float(f"{data['aug']:.2f}")  # AVG

        # Save to memory
        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        # Save into wizard field
        file_data = base64.b64encode(file_stream.read())
        self.write({
            'dummy_file': file_data,
            'dummy_file_name': "staff_analysis.xlsx"
        })

        # Return download action
        return {
            'type': 'ir.actions.act_url',
            'url': f"/web/content/?model=users.wizard&id={self.id}&field=dummy_file&filename_field=dummy_file_name&download=true",
            'target': 'new',  # ensures download in new tab
        }

    def _get_daily_attendance(self, start_date, end_date, unit_name=None):
        """
        Returns a dict:
        {
            user_id: {
                date: {'attendance': 'P/Logged in' or 'A/Not Logged', 'copies': int},
                'total': int,
                'total_copies': int,
                'forms_count': int
            }
        }
        """

        users_domain = [("role", "=", "agent")]
        current_user = request.env.user

        # Handle unit_name logic
        if unit_name == "All" and current_user.role in ["region_head", "circulation_incharge"]:
            if not current_user.exists():
                return {}
            # Get all unit names under this region/circulation head
            if current_user.role == "region_head":
                unit_names = current_user.unit_name_ids.mapped("name")
            else:
                unit_names = [current_user.unit_name] if current_user.unit_name else []
            users_domain.append(("unit_name", "in", unit_names))
        elif unit_name:
            users_domain.append(("unit_name", "=", unit_name))

        # Fetch all agents
        users = self.env["res.users"].sudo().search(users_domain)

        # Fetch sessions and forms
        sessions = self.env["work.session"].sudo().search([
            ('start_time', '>=', start_date),
            ('end_time', '<=', end_date),
        ])
        forms = self.env["customer.form"].sudo().search([
            ('date', '>=', start_date),
            ('date', '<=', end_date),
        ])

        result = {}
        date_range = [
            start_date + timedelta(days=i)
            for i in range((end_date - start_date).days + 1)
        ]

        for user in users:
            result[user.id] = {}
            total_attendance = 0
            total_copies = 0

            for d in date_range:
                # Find sessions for that day
                day_sessions = sessions.filtered(
                    lambda s: s.user_id.id == user.id and s.start_time.date() == d
                )
                attendance_val = "P/Logged in" if day_sessions else "A/Not Logged"
                if day_sessions:
                    total_attendance += 1

                # Count forms that day
                day_forms = forms.filtered(lambda f: f.agent_login == user.login and f.date == d)
                copies_val = len(day_forms)
                total_copies += copies_val

                result[user.id][d] = {
                    "attendance": attendance_val,
                    "copies": copies_val,
                }

            result[user.id]["total"] = total_attendance
            result[user.id]["total_copies"] = total_copies
            result[user.id]["forms_count"] = total_copies  # same for consistency

        return result

    def download_attendance_report(self):
        start_date, end_date = self._get_report_dates()
        # if not start_date or not end_date:
        #     raise UserError("Please select valid period details.")

        unit_name = self.unit_selection  # assuming you store Unit Selection here
        if not unit_name:
            raise UserError("Please select a unit.")
        wb = Workbook()
        ws = wb.active
        ws.title = "Promoters Attendance"

        # --- Header row ---
        ws["A1"].value = "S.No"
        ws["B1"].value = "Name"
        ws.column_dimensions["B"].width = 18

        date_range = [start_date + timedelta(days=i)
                      for i in range((end_date - start_date).days + 1)]

        col = 3  # start column for dates
        for d in date_range:
            left_col = col
            right_col = col + 1

            # Merge date cells
            ws.merge_cells(start_row=1, start_column=left_col, end_row=1, end_column=right_col)

            # Date header
            top_cell = ws.cell(row=1, column=left_col)
            top_cell.value = d.strftime("%d-%m-%Y")
            top_cell.alignment = Alignment(horizontal="center", vertical="center")
            top_cell.font = Font(bold=True, size=10)

            # Sub-headers
            in_cell = ws.cell(row=2, column=left_col)
            out_cell = ws.cell(row=2, column=right_col)
            in_cell.value = "Attendance"
            out_cell.value = "Copies"
            for c in (in_cell, out_cell):
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.font = Font(bold=True, size=9)

            # Column widths
            ws.column_dimensions[get_column_letter(left_col)].width = 12
            ws.column_dimensions[get_column_letter(right_col)].width = 12

            col += 2

        # --- Totals ---
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        ws.cell(row=1, column=col).value = "Total Count"
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=1, column=col).font = Font(bold=True, size=10)
        ws.column_dimensions[get_column_letter(col)].width = 16
        ws.column_dimensions[get_column_letter(col + 1)].width = 16

        in_cell = ws.cell(row=2, column=col)
        out_cell = ws.cell(row=2, column=col + 1)
        in_cell.value = "Total Attendance"
        out_cell.value = "Total Copies"

        # --- UNIT ---
        unit_col = col + 2
        ws.cell(row=1, column=unit_col).value = "UNIT"
        ws.cell(row=1, column=unit_col).alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(unit_col)].width = 14

        # --- Forms Count ---
        forms_col = unit_col + 1
        ws.cell(row=1, column=forms_col).value = "Forms Count"
        ws.cell(row=1, column=forms_col).alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(forms_col)].width = 16

        for c in (in_cell, out_cell):
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = Font(bold=True, size=9)

        # --- Attendance data ---
        data = self._get_daily_attendance(start_date, end_date, unit_name)

        # Fetch users only from the keys in data to ensure consistency
        users = self.env['res.users'].sudo().search([('id', 'in', list(data.keys()))])

        # Group users by unit and sort units by total forms_count descending
        from collections import defaultdict
        unit_groups = defaultdict(list)
        unit_totals = defaultdict(int)
        for user in users:
            unit = user.unit_name or ""
            unit_groups[unit].append(user)
            unit_totals[unit] += data[user.id]['forms_count']

        # Sort units by total forms descending
        sorted_units = sorted(unit_totals.keys(), key=lambda u: unit_totals[u], reverse=True)

        # For each unit, sort users by total_copies descending
        sorted_users = []
        for unit in sorted_units:
            unit_users = sorted(unit_groups[unit], key=lambda u: data[u.id]['total_copies'], reverse=True)
            sorted_users.extend(unit_users)

        row = 3
        for idx, user in enumerate(sorted_users, start=1):
            ws.cell(row=row, column=1).value = idx
            ws.cell(row=row, column=2).value = user.name

            col_ptr = 3
            for d in date_range:
                ws.cell(row=row, column=col_ptr).value = data[user.id][d]["attendance"]  # Attendance
                ws.cell(row=row, column=col_ptr + 1).value = data[user.id][d]["copies"]  # Copies
                col_ptr += 2

            ws.cell(row=row, column=col_ptr).value = data[user.id]['total']
            ws.cell(row=row, column=col_ptr + 1).value = data[user.id]['total_copies']
            ws.cell(row=row, column=unit_col).value = user.unit_name or ""
            ws.cell(row=row, column=forms_col).value = data[user.id]['forms_count']

            row += 1

        # --- Save file ---
        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        file_data = base64.b64encode(file_stream.read())

        self.write({
            'dummy_file': file_data,
            'dummy_file_name': "attendance_report.xlsx"
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f"/web/content/?model=users.wizard&id={self.id}&field=dummy_file&filename_field=dummy_file_name&download=true",
            'target': 'new',
        }